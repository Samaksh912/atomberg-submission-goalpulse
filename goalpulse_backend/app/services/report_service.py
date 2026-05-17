"""Report generation service — CSV and Excel achievement reports."""

from __future__ import annotations

import csv
import io
from typing import Any, Optional

from app.services.firebase_service import db
from app.services.analytics_service import _goals_for, _checkins_for_goals, _user_map

GOALS = "goals"
CHECKINS = "checkins"


def _build_rows(
    requester_id: str,
    role: str,
    cycle_id: str,
    quarter: Optional[str],
) -> list[dict[str, Any]]:
    """Build the flat list of achievement rows."""
    goals = _goals_for(requester_id, role, cycle_id)
    users = _user_map()
    goal_ids = [g["id"] for g in goals]
    checkins = _checkins_for_goals(goal_ids)

    # Index: goal_id → list[checkin]
    ci_by_goal: dict[str, list[dict]] = {}
    for c in checkins:
        ci_by_goal.setdefault(c["goal_id"], []).append(c)

    rows: list[dict[str, Any]] = []
    quarters = [quarter] if quarter else ["Q1", "Q2", "Q3", "Q4"]

    for g in goals:
        uid = g.get("employee_id", "")
        user = users.get(uid, {})
        emp_name = user.get("display_name") or user.get("email", uid)
        dept = user.get("department", "")
        sheet_status = g.get("sheet_status", "draft")

        goal_items = {gi["goal_item_id"]: gi for gi in g.get("goals", [])}

        for q in quarters:
            # Find matching checkin.
            checkin = next(
                (c for c in ci_by_goal.get(g["id"], []) if c.get("quarter") == q),
                None,
            )

            for item_id, item in goal_items.items():
                # Get actual entry from checkin.
                actual_entry: Optional[dict] = None
                progress_score = 0.0
                actual_val = None

                if checkin:
                    for a in checkin.get("actuals", []):
                        if a.get("goal_item_id") == item_id:
                            actual_entry = a
                            progress_score = float(a.get("progress_score", 0))
                            actual_val = a.get("actual")
                            break

                rows.append(
                    {
                        "employee_name": emp_name,
                        "department": dept,
                        "quarter": q,
                        "goal_title": item.get("title", ""),
                        "thrust_area": item.get("thrust_area", ""),
                        "uom_type": item.get("uom_type", ""),
                        "planned_target": item.get("target", ""),
                        "actual_achievement": actual_val if actual_val is not None else "—",
                        "progress_score": round(progress_score, 1),
                        "weightage": item.get("weightage", 0),
                        "status": sheet_status,
                        "checkin_submitted": "Yes" if checkin else "No",
                    }
                )

    return rows


class ReportService:
    """Generate achievement reports in various formats."""

    COLUMNS = [
        "employee_name",
        "department",
        "quarter",
        "goal_title",
        "thrust_area",
        "uom_type",
        "planned_target",
        "actual_achievement",
        "progress_score",
        "weightage",
        "status",
        "checkin_submitted",
    ]

    HEADERS = [
        "Employee Name",
        "Department",
        "Quarter",
        "Goal Title",
        "Thrust Area",
        "UoM Type",
        "Planned Target",
        "Actual Achievement",
        "Progress Score (%)",
        "Weightage (%)",
        "Sheet Status",
        "Check-In Submitted",
    ]

    @staticmethod
    def generate_achievement_report(
        requester_id: str,
        role: str,
        cycle_id: str,
        quarter: Optional[str] = None,
        format: str = "json",
    ):
        """Generate report in json, csv, or excel format."""
        rows = _build_rows(requester_id, role, cycle_id, quarter)

        if format == "json":
            return rows

        if format == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(ReportService.HEADERS)
            for row in rows:
                writer.writerow([row.get(c, "") for c in ReportService.COLUMNS])
            buf.seek(0)
            return buf

        if format == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                # Fallback: return CSV as bytes.
                buf = io.StringIO()
                csv.writer(buf).writerow(ReportService.HEADERS)
                for row in rows:
                    csv.writer(buf).writerow(
                        [row.get(c, "") for c in ReportService.COLUMNS]
                    )
                buf.seek(0)
                return io.BytesIO(buf.read().encode("utf-8"))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Achievement Report"

            # Header styling.
            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(bold=True, color="FFFFFF")
            for col_idx, h in enumerate(ReportService.HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, 2):
                for col_idx, col in enumerate(ReportService.COLUMNS, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))

            # Auto-width (approx).
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        return rows


report_service = ReportService()
